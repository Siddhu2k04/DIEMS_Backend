# pyrefly: ignore [missing-import]
from flask import Blueprint, jsonify, request, send_file
# pyrefly: ignore [missing-import]
from flask_jwt_extended import jwt_required, get_jwt_identity
from models import Event, User, Registration, Feedback
from extensions import db
from sqlalchemy import func
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

organizer_bp = Blueprint('organizer', __name__)

@organizer_bp.route('/events', methods=['GET'])
@jwt_required()
def get_organizer_events():
    current_user_id = get_jwt_identity()
    user = User.query.get(current_user_id)
    
    if user.role not in ['organizer', 'admin']:
        return jsonify({"msg": "Unauthorized."}), 403

    events = Event.query.filter_by(organizer_id=current_user_id).all()
    result = []
    for event in events:
        result.append({
            "id": event.id,
            "title": event.title,
            "status": event.status,
            "date_time": event.date_time.isoformat(),
            "venue": event.venue,
            "category": event.category,
            "registration_limit": event.registration_limit,
            "registration_count": Registration.query.filter_by(event_id=event.id).count()
        })
    return jsonify(result), 200

@organizer_bp.route('/events/<int:event_id>/registrations', methods=['GET'])
@jwt_required()
def get_event_registrations(event_id):
    current_user_id = get_jwt_identity()
    user = User.query.get(current_user_id)
    event = Event.query.get_or_404(event_id)

    if user.role != 'admin' and event.organizer_id != int(current_user_id):
        return jsonify({"msg": "Unauthorized."}), 403

    registrations = Registration.query.filter_by(event_id=event_id).all()
    result = []
    for reg in registrations:
        result.append({
            "id": reg.id,
            "user_name": reg.user.name,
            "user_email": reg.user.email,
            "status": reg.status,
            "attendance_status": reg.attendance_status,
            "registered_at": reg.registered_at.isoformat()
        })
    return jsonify(result), 200


@organizer_bp.route('/analytics', methods=['GET'])
@jwt_required()
def get_organizer_analytics():
    current_user_id = get_jwt_identity()
    user = User.query.get(current_user_id)

    if user.role not in ['organizer', 'admin']:
        return jsonify({"msg": "Unauthorized."}), 403

    events = Event.query.filter_by(organizer_id=current_user_id).all()
    total_events = len(events)

    # total registrations for organizer's events
    total_registrations = Registration.query.join(Event).filter(Event.organizer_id == current_user_id).count()

    # average rating across organizer's events
    avg_rating = db.session.query(func.avg(Feedback.rating)).join(Event, Feedback.event_id == Event.id).filter(Event.organizer_id == current_user_id).scalar()
    avg_rating = round(float(avg_rating), 2) if avg_rating is not None else None

    # conversion rate = total_registrations / total_capacity (only events with registration_limit)
    total_capacity = sum([e.registration_limit for e in events if e.registration_limit])
    conversion_rate = None
    if total_capacity and total_capacity > 0:
        conversion_rate = round((total_registrations / total_capacity) * 100, 2)

    # registrations per event summary
    events_summary = []
    for e in events:
        count = Registration.query.filter_by(event_id=e.id).count()
        events_summary.append({
            "id": e.id,
            "title": e.title,
            "registration_count": count,
            "registration_limit": e.registration_limit
        })

    return jsonify({
        "total_events": total_events,
        "total_registrations": total_registrations,
        "avg_rating": avg_rating,
        "conversion_rate": conversion_rate,
        "events": events_summary
    }), 200

@organizer_bp.route('/registrations/<int:reg_id>/status', methods=['PUT'])
@jwt_required()
def update_registration_status(reg_id):
    current_user_id = get_jwt_identity()
    user = User.query.get(current_user_id)
    registration = Registration.query.get_or_404(reg_id)
    event = registration.event

    if user.role != 'admin' and event.organizer_id != int(current_user_id):
        return jsonify({"msg": "Unauthorized."}), 403

    data = request.get_json()
    new_status = data.get('status')
    if new_status in ['pending', 'approved', 'rejected']:
        registration.status = new_status
        db.session.commit()
        return jsonify({"msg": f"Registration marked as {new_status}"}), 200
    
    return jsonify({"msg": "Invalid status"}), 400

@organizer_bp.route('/registrations/<int:reg_id>/attendance', methods=['PUT'])
@jwt_required()
def update_attendance(reg_id):
    current_user_id = get_jwt_identity()
    user = User.query.get(current_user_id)
    registration = Registration.query.get_or_404(reg_id)
    event = registration.event

    if user.role != 'admin' and event.organizer_id != int(current_user_id):
        return jsonify({"msg": "Unauthorized."}), 403

    data = request.get_json() or {}
    registration.attendance_status = bool(data.get('attendance_status', True))
    db.session.commit()
    return jsonify({
        "msg": "Attendance updated",
        "attendance_status": registration.attendance_status
    }), 200

@organizer_bp.route('/events/<int:event_id>/registrations/download', methods=['GET'])
@jwt_required()
def download_registrations_excel(event_id):
    current_user_id = get_jwt_identity()
    user = User.query.get(current_user_id)
    event = Event.query.get_or_404(event_id)

    if user.role != 'admin' and event.organizer_id != int(current_user_id):
        return jsonify({"msg": "Unauthorized."}), 403

    registrations = Registration.query.filter_by(event_id=event_id).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Participants"

    # -- Styles --
    header_font = Font(name='Calibri', bold=True, size=14, color='FFFFFF')
    header_fill = PatternFill(start_color='4F46E5', end_color='4F46E5', fill_type='solid')
    col_header_font = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
    col_header_fill = PatternFill(start_color='6366F1', end_color='6366F1', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin', color='D1D5DB'),
        right=Side(style='thin', color='D1D5DB'),
        top=Side(style='thin', color='D1D5DB'),
        bottom=Side(style='thin', color='D1D5DB')
    )
    center_align = Alignment(horizontal='center', vertical='center')

    # -- Event title header row --
    ws.merge_cells('A1:F1')
    title_cell = ws['A1']
    title_cell.value = f"Participant Records — {event.title}"
    title_cell.font = header_font
    title_cell.fill = header_fill
    title_cell.alignment = center_align
    ws.row_dimensions[1].height = 30

    # -- Event info row --
    ws.merge_cells('A2:F2')
    info_cell = ws['A2']
    event_date = event.date_time.strftime('%d %b %Y, %I:%M %p') if event.date_time else 'N/A'
    info_cell.value = f"Venue: {event.venue}  |  Date: {event_date}  |  Total Participants: {len(registrations)}"
    info_cell.font = Font(name='Calibri', size=10, italic=True, color='374151')
    info_cell.alignment = center_align
    ws.row_dimensions[2].height = 22

    # -- Blank spacer row --
    ws.row_dimensions[3].height = 8

    # -- Column headers (row 4) --
    columns = ['Sr. No', 'Name', 'Email', 'Department', 'Registration Status']
    for col_num, col_title in enumerate(columns, 1):
        cell = ws.cell(row=4, column=col_num, value=col_title)
        cell.font = col_header_font
        cell.fill = col_header_fill
        cell.alignment = center_align
        cell.border = thin_border
    ws.row_dimensions[4].height = 24

    # -- Data rows --
    for idx, reg in enumerate(registrations, 1):
        participant = User.query.get(reg.user_id)
        row_num = idx + 4
        row_data = [
            idx,
            participant.name if participant else 'N/A',
            participant.email if participant else 'N/A',
            participant.department if participant and participant.department else 'N/A',
            reg.status.capitalize() if reg.status else 'N/A'
        ]
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.border = thin_border
            cell.alignment = center_align
            cell.font = Font(name='Calibri', size=10)

        # Alternate row shading
        if idx % 2 == 0:
            for col_num in range(1, len(columns) + 1):
                ws.cell(row=row_num, column=col_num).fill = PatternFill(
                    start_color='F3F4F6', end_color='F3F4F6', fill_type='solid'
                )

    # -- Auto-size columns --
    for col_num, col_title in enumerate(columns, 1):
        max_length = len(col_title)
        for row in range(5, len(registrations) + 5):
            cell_value = ws.cell(row=row, column=col_num).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))
        ws.column_dimensions[ws.cell(row=4, column=col_num).column_letter].width = max_length + 4

    # -- Write to buffer and send --
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    safe_title = "".join(c if c.isalnum() or c in (' ', '-', '_') else '_' for c in event.title)
    filename = f"{safe_title}_participants.xlsx"

    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
